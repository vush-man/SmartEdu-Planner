from io import BytesIO
from .models import Teacher, Lecture, db, User
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Spacer

def find_and_assign_substitute(lecture_id: int):
    """
    Finds and assigns a substitute for a given lecture.
    Returns a tuple of (new_teacher, message).
    """
    lecture = Lecture.query.get(lecture_id)
    if not lecture:
        return None, "Error: Lecture not found."

    original_teacher = lecture.teacher

    potential_substitutes = Teacher.query.filter(
        Teacher.subject == lecture.subject,
        Teacher.id != original_teacher.id
    ).all()

    for substitute in potential_substitutes:
        conflict = Lecture.query.filter_by(
            teacher_id=substitute.id,
            time_slot=lecture.time_slot,
            day=lecture.day
        ).first()

        if not conflict:
            lecture.teacher_id = substitute.id
            lecture.status = "Reassigned"
            db.session.commit()
            message = f"Success: {substitute.name} assigned to lecture {lecture.id} for section {lecture.section}."
            return substitute, message

    lecture.status = "Cancelled"
    db.session.commit()
    return None, f"No substitute available for lecture {lecture.id}. Lecture has been cancelled."


def generate_timetable_pdf(user):
    """Generates a grid-based PDF timetable for a given student user."""
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(letter))
    
    section = user.section
    lectures = Lecture.query.filter_by(section=section).all()

    days = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday"]
    time_slots = sorted(list(set(l.time_slot for l in lectures)))

    data = [['Time Slot'] + days]

    timetable = {time: {day: "" for day in days} for time in time_slots}
    for lecture in lectures:
        cell_text = f"{lecture.subject}\n({lecture.teacher.name})"
        timetable[lecture.time_slot][lecture.day] = cell_text

    for time in time_slots:
        row = [time] + [timetable[time][day] for day in days]
        data.append(row)

    table = Table(data)
    style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('MIN-HEIGHT', (0, 0), (-1, -1), 40),
    ])
    table.setStyle(style)

    title = f"Timetable for Section {section} (Student: {user.username})"
    elements = [Table([[title]], style=[('ALIGN', (0, 0), (-1, -1), 'CENTER'), ('FONTSIZE', (0,0), (-1,-1), 16)])]

    elements.append(Spacer(1, 20))
    
    elements.append(table)
    
    doc.build(elements)
    
    return buffer


def generate_timetable_excel(user):
    """Generates a styled Excel timetable for a given student user."""
    section = user.section
    lectures = Lecture.query.filter_by(section=section).all()
    
    wb = Workbook()
    ws = wb.active
    ws.title = f"Timetable Section {section}"

    days = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday"]
    time_slots = sorted(list(set(l.time_slot for l in lectures)))

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    ws.merge_cells('A1:G1')
    title_cell = ws['A1']
    title_cell.value = f"Timetable for Section {section} (Student: {user.username})"
    title_cell.font = Font(size=16, bold=True)
    title_cell.alignment = Alignment(horizontal='center')

    ws.append([''] + days)
    for col_idx, day in enumerate(days, 2):
        cell = ws.cell(row=2, column=col_idx, value=day)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = border
        ws.column_dimensions[chr(65 + col_idx -1)].width = 20

    ws.column_dimensions['A'].width = 15
    ws.cell(row=2, column=1).font = header_font
    ws.cell(row=2, column=1).fill = header_fill

    timetable = {time: {day: "" for day in days} for time in time_slots}
    for lecture in lectures:
        timetable[lecture.time_slot][lecture.day] = f"{lecture.subject}\n({lecture.teacher.name})"
        
    for row_idx, time in enumerate(time_slots, 3):
        row_data = [time] + [timetable[time][day] for day in days]
        ws.append(row_data)
        for col_idx in range(1, len(row_data) + 1):
             cell = ws.cell(row=row_idx, column=col_idx)
             cell.alignment = center_align
             cell.border = border

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return buffer
